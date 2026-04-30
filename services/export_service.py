import io
from typing import List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from config import settings


class ExportService:
    """Excel导出服务"""

    @staticmethod
    def save_to_excel_bytes(data_list: List[List[str]]) -> bytes:
        """保存数据到Excel并返回bytes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "B站搜索结果"

        # 添加表头
        ws.append(settings.HEADERS)

        # 添加数据
        for row in data_list:
            safe_row = [str(cell) if cell is not None else "N/A" for cell in row]
            ws.append(safe_row)

        # 动态计算列宽
        for idx, header in enumerate(settings.HEADERS, 1):
            col_letter = get_column_letter(idx)
            max_length = max(len(str(header)),
                             max((len(str(row[idx - 1])) for row in data_list), default=0))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # 保存到bytes
        #made by github-suslock
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.read()